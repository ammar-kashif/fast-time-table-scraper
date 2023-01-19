# import openpyxl and pandas
import openpyxl as op
from pandas import *

# open the workbook
wb = op.load_workbook('FSCTimeTable.xlsx')

# get the sheet
sheet = wb['Monday']

# dictionary of subject colours
subjCol = {
    "BS CS (2022)": "FFFFB740",
    "BS CS (2021)": "FF7F6000",
    "BS CS (2020)": "FFF1C232",
    "BS CS (2019)": "FFFFE599",
    "BS DS (2022)": "FF7F4CFF",
    "BS DS (2021)": "FF351C75",
    "BS DS (2020)": "FFB17FD7",
    "BS DS (2019)": "FFB4A7D6",
    "BS AI (2022)": "FF00F600",
    "BS AI (2021)": "FF274E13",
    "BS AI (2020)": "FF6AA84F",
    "BS AI (2019)": "FFB6D7A8",
    "BS SE (2022)": "FFE62C06",
    "BS SE (2021)": "FF85200C",
    "BS SE (2020)": "FFDD7E6B",
    "BS SE (2019)": "FFF4CCCC",
    "BS CY (2022)": "0000FF00",
    "BS CY (2021)": "FF0050EF",
    "BS CY (2020)": "FF2B8CC4",
    "BS CY (2019)": "FF92CDDC",
}

# user inputs
batch = "BS DS (2022)"
section = "D"
subSection = "D2"
color = subjCol[batch]
# slice "DS" from BS DS (2022) and store it in section
section = batch[3:5] + "-" + section
subSection = batch[3:5] + "-" + subSection


# list of timings
schedule = [[]]

# loop over 52 rows and 12 columns
for i in range(1, 55):
    for j in range(1, 13):
        # get the cell
        cell = sheet.cell(row=i, column=j)
        # get the cell value if it is not empty
        if cell.value is not None:
            # if ":" in cell.value:
            #     print(cell.value)
            #     print(i, j)
            # get the cell value if it has the colour code of batch
            if cell.fill.start_color.index == color:
                # get the cell value if the cell value contains the section or subSection
                if section in cell.value or subSection in cell.value:
                    # getting timings 
                    if ":" in cell.value:
                        subject = cell.value[:-11]
                        room = sheet.cell(row=i, column=1)
                        timings = cell.value[-11:]
                        schedule.append([room, subject, timings])
                    else:
                        subject = cell.value
                        room = sheet.cell(row=i, column=1) # masla
                        timings = sheet.cell(row=38, column=j)  
                        schedule.append([room, subject, timings])

# print the schedule
print(DataFrame(schedule, columns=["Room", "Subject", "Timings"]))