# import openpyxl and pandas
import openpyxl as op
import pandas as pd
from pandas import *

# open the workbook
wb = op.load_workbook('Ramadan-TimeTable-FSC-Spring-2023.xlsx')

# get the sheet
sheetMonday = wb['Monday']
sheetTuesday = wb['Tuesday']
sheetWednesday = wb['Wednesday (Apr. 12, 2023)']
sheetThursday = wb['Thursday (Apr. 13, 2023)']
sheetFriday = wb['Friday']

# dictionary of subject colours
batchYear = {
    "FFFFB740": "2022",
    "FF7F6000": "2021",
    "FFF1C232": "2020",
    "FFFFE599": "2019",
    "FF7F4CFF": "2022",
    "FF351C75": "2021",
    "FFB17FD7": "2020",
    "FFB4A7D6": "2019",
    "FF00F600": "2022",
    "FF274E13": "2021",
    "FF6AA84F": "2020",
    "FFB6D7A8": "2019",
    "FFE62C06": "2022",
    "FF85200C": "2021",
    "FFDD7E6B": "2020",
    "FFF4CCCC": "2019",
    "FF0000FF": "2022",
    "FF0050EF": "2021",
    "FF599DDA": "2020",
    "FFABCCEB": "2019"
}

# dictionary of subject colours
departments = ["DS", "CS", "AI", "SE", "CY"]

# list to store the schedule
schedule = [[]]

# print(sheet.cell(row=22, column=10).value)
# print(sheet.cell(row=22, column=10).fill.start_color.index)

# Monday
for row in range(2, 55):
    for col in range(2, 14):
        # get the cell
        cell = sheetMonday.cell(row=row, column=col)
        # get the cell value if it is not empty
        if cell.value is None:
            continue

        valid = False
        for dep in departments:
            if dep not in cell.value or ")" not in cell.value:
                continue
            else:
                dep = dep
                valid = True
                break

        if not valid:
            continue
        
        if batchYear.get(cell.fill.start_color.index) is not None:
            batch = batchYear[cell.fill.start_color.index]

        if ")" and "-" in cell.value:
            if  ":" not in cell.value:
                spliced = cell.value.split("-")
                spliced = spliced[-1].split(")")
                section = spliced[0]
            else:
                remTiming = cell.value[:-11]
                spliced = remTiming.split("-")
                spliced = spliced[-1].split(")")
                section = spliced[0]

        if ":" in cell.value:
            subject = cell.value[:-11]
            room = sheetMonday.cell(row=row, column=1).value
            timings = cell.value[-11:]
        else:
            subject = cell.value
            room = sheetMonday.cell(row=row, column=1).value
            if "Lab" in cell.value:
                timings = sheetMonday.cell(row=38, column=col).value
            else:
                timings = sheetMonday.cell(row=1, column=col).value

        schedule.append(["Monday", dep, batch, section, subject, room, timings])

# Tuesday
for row in range(2, 55):
    for col in range(2, 14):
        # get the cell
        cell = sheetTuesday.cell(row=row, column=col)
        # get the cell value if it is not empty
        if cell.value is None:
            continue

        valid = False
        for dep in departments:
            if dep not in cell.value or ")" not in cell.value:
                continue
            else:
                dep = dep
                valid = True
                break

        if not valid:
            continue
        
        if batchYear.get(cell.fill.start_color.index) is not None:
            batch = batchYear[cell.fill.start_color.index]

        if ")" and "-" in cell.value:
            if  ":" not in cell.value:
                spliced = cell.value.split("-")
                spliced = spliced[-1].split(")")
                section = spliced[0]
            else:
                remTiming = cell.value[:-11]
                spliced = remTiming.split("-")
                spliced = spliced[-1].split(")")
                section = spliced[0]

        if ":" in cell.value:
            subject = cell.value[:-11]
            room = sheetTuesday.cell(row=row, column=1).value
            timings = cell.value[-11:]
        else:
            subject = cell.value
            room = sheetTuesday.cell(row=row, column=1).value
            if "Lab" in cell.value:
                timings = sheetTuesday.cell(row=38, column=col).value
            else:
                timings = sheetTuesday.cell(row=1, column=col).value

        schedule.append(["Tuesday", dep, batch, section, subject, room, timings])

# Wednesday
for row in range(2, 55):
    for col in range(2, 14):
        # get the cell
        cell = sheetWednesday.cell(row=row, column=col)
        # get the cell value if it is not empty
        if cell.value is None:
            continue

        valid = False
        for dep in departments:
            if dep not in cell.value or ")" not in cell.value:
                continue
            else:
                dep = dep
                valid = True
                break

        if not valid:
            continue
        
        if batchYear.get(cell.fill.start_color.index) is not None:
            batch = batchYear[cell.fill.start_color.index]

        if ")" and "-" in cell.value:
            if  ":" not in cell.value:
                spliced = cell.value.split("-")
                spliced = spliced[-1].split(")")
                section = spliced[0]
            else:
                remTiming = cell.value[:-11]
                spliced = remTiming.split("-")
                spliced = spliced[-1].split(")")
                section = spliced[0]

        if ":" in cell.value:
            subject = cell.value[:-11]
            room = sheetWednesday.cell(row=row, column=1).value
            timings = cell.value[-11:]
        else:
            subject = cell.value
            room = sheetWednesday.cell(row=row, column=1).value
            if "Lab" in cell.value:
                timings = sheetWednesday.cell(row=38, column=col).value
            else:
                timings = sheetWednesday.cell(row=1, column=col).value

        schedule.append(["Wednesday", dep, batch, section, subject, room, timings])

# Thursday
for row in range(2, 55):
    for col in range(2, 14):
        # get the cell
        cell = sheetThursday.cell(row=row, column=col)
        # get the cell value if it is not empty
        if cell.value is None:
            continue

        valid = False
        for dep in departments:
            if dep not in cell.value or ")" not in cell.value:
                continue
            else:
                dep = dep
                valid = True
                break

        if not valid:
            continue
        
        if batchYear.get(cell.fill.start_color.index) is not None:
            batch = batchYear[cell.fill.start_color.index]

        if ")" and "-" in cell.value:
            if  ":" not in cell.value:
                spliced = cell.value.split("-")
                spliced = spliced[-1].split(")")
                section = spliced[0]
            else:
                remTiming = cell.value[:-11]
                spliced = remTiming.split("-")
                spliced = spliced[-1].split(")")
                section = spliced[0]

        if ":" in cell.value:
            subject = cell.value[:-11]
            room = sheetThursday.cell(row=row, column=1).value
            timings = cell.value[-11:]
        else:
            subject = cell.value
            room = sheetThursday.cell(row=row, column=1).value
            if "Lab" in cell.value:
                timings = sheetThursday.cell(row=38, column=col).value
            else:
                timings = sheetThursday.cell(row=1, column=col).value

        schedule.append(["Thursday", dep, batch, section, subject, room, timings])

# Friday
for row in range(2, 55):
    for col in range(2, 14):
        # get the cell
        cell = sheetFriday.cell(row=row, column=col)
        # get the cell value if it is not empty
        if cell.value is None:
            continue

        valid = False
        for dep in departments:
            if dep not in cell.value or ")" not in cell.value:
                continue
            else:
                dep = dep
                valid = True
                break

        if not valid:
            continue
        
        if batchYear.get(cell.fill.start_color.index) is not None:
            batch = batchYear[cell.fill.start_color.index]

        if ")" and "-" in cell.value:
            if  ":" not in cell.value:
                spliced = cell.value.split("-")
                spliced = spliced[-1].split(")")
                section = spliced[0]
            else:
                remTiming = cell.value[:-11]
                spliced = remTiming.split("-")
                spliced = spliced[-1].split(")")
                section = spliced[0]

        if ":" in cell.value:
            subject = cell.value[:-11]
            room = sheetFriday.cell(row=row, column=1).value
            timings = cell.value[-11:]
        else:
            subject = cell.value
            room = sheetFriday.cell(row=row, column=1).value
            if "Lab" in cell.value:
                timings = sheetFriday.cell(row=38, column=col).value
            else:
                timings = sheetFriday.cell(row=1, column=col).value

        schedule.append(["Friday", dep, batch, section, subject, room, timings])

# print the schedule
print(DataFrame(schedule[1:], columns=["Day", "Department", "Batch", "Section", "Subject", "Room", "Timings"]))
df = pd.DataFrame(schedule[1:], columns=["Day", "Department", "Batch", "Section", "Subject", "Room", "Timings"])

df.to_json(r'C:\Programming\Personal\fast-time-table\timetable.json', orient='records')